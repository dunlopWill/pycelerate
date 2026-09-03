import pytest

from pycelerate import F, Sheet, cell

openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture
def wb():
    book = openpyxl.Workbook()
    book.active.title = "Model"
    return book


@pytest.fixture
def s(wb):
    return Sheet(wb.active)


def test_put_writes_the_value_and_returns_its_reference(s):
    rev = s.put("B2", 1000)
    assert s.ws["B2"].value == 1000
    assert rev == cell("B2")
    assert (rev.row, rev.col) == (2, 2)


def test_put_by_position(s):
    rev = s.put(2, 2, 1000)
    assert s.ws["B2"].value == 1000
    assert str(rev) == "=B2"


def test_the_coordinate_is_never_typed_twice(s):
    rev, cost = s.put("B2", 1000), s.put("B3", 600)
    margin = s.put("B4", (rev - cost) / rev)
    assert s.ws["B4"].value == "=(B2-B3)/B2"
    # ...and the returned ref composes further.
    s.put("B5", F.ROUND(margin * 100, 1))
    assert s.ws["B5"].value == "=ROUND(B4*100,1)"


def test_same_sheet_references_stay_unqualified(s):
    rev, cost = s.put("B2", 1000), s.put("B3", 600)
    assert str(rev - cost) == "=B2-B3"  # not "=Model!B2-Model!B3"


def test_crossing_sheets_supplies_the_name_for_you(wb):
    src = Sheet(wb.active)
    rev = src.put("B2", 1000)
    out = wb.create_sheet("Summary")
    out["A1"] = rev.qualified() * 2
    assert out["A1"].value == "=Model!B2*2"
    assert str(rev.qualified().bare()) == "=B2"


def test_qualified_without_a_home_is_an_error():
    with pytest.raises(ValueError, match="no home sheet"):
        cell("B2").qualified()


def test_styling_passes_through(s):
    from openpyxl.styles import Font

    ref = s.put("B2", 1000, number_format="#,##0", font=Font(bold=True))
    assert s.ws["B2"].number_format == "#,##0"
    assert s.ws["B2"].font.bold is True
    assert str(ref) == "=B2"


def test_references_without_writing(s):
    assert str(s["B2"]) == "=B2"
    assert str(s["A1:A9"]) == "=A1:A9"
    assert str(s.cell(2, 2)) == "=B2"
    assert s["B2"].home == "Model"
    assert str(s["A1:A9"].qualified()) == "=Model!A1:A9"


def test_span_and_column(s):
    top = s.put("B2", 1)
    bottom = s.put("B9", 8)
    assert str(s.span(top, bottom).sum()) == "=SUM(B2:B9)"
    assert str(s.column(top)) == "=B:B"
    assert str(s.column("D")) == "=D:D"


def test_ws_escape_hatch_still_works(s):
    s.ws["Z1"] = "plain openpyxl"
    assert s.ws["Z1"].value == "plain openpyxl"
    assert s.title == "Model"


@pytest.mark.parametrize("args", [("B2",), (1, 2), ("B2", 1, 2), ()])
def test_bad_put_signatures(s, args):
    with pytest.raises(TypeError):
        s.put(*args)


def test_roundtrip_through_a_saved_file(wb, tmp_path):
    s = Sheet(wb.active)
    rev, cost = s.put("B2", 1000), s.put("B3", 600)
    s.put("B4", F.IF(rev.gt(cost), rev - cost, 0))

    path = tmp_path / "book.xlsx"
    wb.save(path)
    reloaded = openpyxl.load_workbook(path)["Model"]
    assert reloaded["B4"].value == "=IF(B2>B3,B2-B3,0)"
    assert reloaded["B4"].data_type == "f"
    assert reloaded["B2"].value == 1000
